from .shared import *

from dataclasses import asdict
import re

from med_autoscience.literature_records import LiteratureRecord


def _open_submission_route_context() -> dict[str, object]:
    return {
        "authority_snapshot": {
            "surface": "authority_snapshot",
            "dispatch_gate": {
                "state": "open",
                "dispatch_allowed": True,
                "blocking_reasons": [],
            },
            "route_authorization": {
                "authorized": True,
                "paper_write_allowed": True,
                "bundle_build_allowed": True,
                "runtime_recovery_allowed": True,
            },
            "authority_refs": {
                "study_truth": {"epoch": "truth-1"},
                "runtime_health": {"epoch": "runtime-1"},
            },
        }
    }


def test_inspect_submission_source_markdown_counts_short_f_headings_as_figures(tmp_path: Path) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    source_markdown = tmp_path / "manuscript_submission.md"
    write_text(
        source_markdown,
        """---
title: "Submission Manuscript"
---

# Figures

## F1. Main figure

![](figures/F1_main.png)

Legend text for the short-F main figure.
""",
    )

    inspection = module.inspect_submission_source_markdown(source_markdown)

    assert inspection["figure_block_count"] == 1
    assert inspection["figure_blocks_with_images"] == 1
    assert inspection["figure_blocks_with_legends"] == 1


def test_inspect_submission_source_markdown_counts_independent_figure_legends_section(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    source_markdown = tmp_path / "manuscript_submission.md"
    write_text(
        source_markdown,
        """---
title: "Submission Manuscript"
---

# Figures

## Figure 1. Main figure

![](figures/F1_main.png)

# Figure Legends

Figure 1. Legend text for the independent figure legend section.

# Tables

## Table 1

| Characteristic | Value |
| --- | --- |
| Age | 52 |
""",
    )

    inspection = module.inspect_submission_source_markdown(source_markdown)

    assert inspection["figure_block_count"] == 1
    assert inspection["figure_blocks_with_images"] == 1
    assert inspection["figure_blocks_with_legends"] == 1


def test_inspect_submission_source_markdown_treats_lowercase_figure_legends_as_separate_section(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    source_markdown = tmp_path / "manuscript_submission.md"
    write_text(
        source_markdown,
        """---
title: "Submission Manuscript"
---

# Figures

## Figure 1. Main figure

![](figures/F1_main.png)

# Figure legends

## Figure 1. Main figure

Legend text for the lowercase independent figure legend section.

# Tables

## Table 1

| Characteristic | Value |
| --- | --- |
| Age | 52 |
""",
    )

    inspection = module.inspect_submission_source_markdown(source_markdown)

    assert inspection["figure_block_count"] == 1
    assert inspection["figure_blocks_with_images"] == 1
    assert inspection["figure_blocks_with_legends"] == 1


def test_create_submission_minimal_package_supports_manuscript_shaped_draft_without_front_matter(
    tmp_path: Path,
    real_submission_exports,
) -> None:
    from docx import Document

    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_root = paper_root / "submission_minimal"
    compiled_submission_markdown = submission_root / "manuscript_submission.md"
    assert compiled_submission_markdown.exists()

    submission_markdown = compiled_submission_markdown.read_text(encoding="utf-8")
    assert submission_markdown.startswith("---\n")
    assert 'title: "Manuscript-Shaped Draft Title"' in submission_markdown
    assert "bibliography: references.bib" in submission_markdown
    assert "bibliography: ../../references.bib" not in submission_markdown
    assert "\n# References\n" in submission_markdown
    assert "title: \"Article Title\"" not in submission_markdown
    assert "\n# Methods\n\nStudy methods paragraph.\n" in submission_markdown
    assert "Draft abstract methods." in submission_markdown
    assert manifest["manuscript"]["source_markdown_path"] == "paper/submission_minimal/manuscript_submission.md"

    document = Document(submission_root / "manuscript.docx")
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    assert paragraphs[0] == "Manuscript-Shaped Draft Title"
    assert any("Study methods paragraph." in paragraph for paragraph in paragraphs)
    assert any("A primary source" in paragraph for paragraph in paragraphs)


def test_create_submission_minimal_package_supports_front_matter_manuscript_shaped_draft(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    write_text(
        paper_root / "draft.md",
        """---
title: "Front Matter Manuscript-Shaped Draft"
bibliography: references.bib
---

## Abstract

Structured abstract paragraph.

## Introduction

Frontmatter introduction paragraph.

## Methods

Frontmatter methods paragraph.

## Results

Frontmatter results paragraph.

## Discussion

Frontmatter discussion paragraph.
""",
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(
        encoding="utf-8"
    )
    assert 'title: "Front Matter Manuscript-Shaped Draft"' in submission_markdown
    for heading in ["# Abstract", "# Introduction", "# Methods", "# Results", "# Discussion"]:
        assert submission_markdown.splitlines().count(heading) == 1
    for paragraph in [
        "Structured abstract paragraph.",
        "Frontmatter introduction paragraph.",
        "Frontmatter methods paragraph.",
        "Frontmatter results paragraph.",
        "Frontmatter discussion paragraph.",
    ]:
        assert submission_markdown.count(paragraph) == 1


def test_create_submission_minimal_package_keeps_submission_figure_legends_concise(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_paper_workspace(tmp_path)
    dump_json(
        paper_root / "figure_semantics_manifest.json",
        {
            "schema_version": 1,
            "figures": [
                {
                    "figure_id": "F1",
                    "direct_message": "The cohort architecture is summarized for reader orientation.",
                    "clinical_implication": "The figure summarizes the observed cohort architecture.",
                    "interpretation_boundary": "Do not recast this figure as causal proof.",
                    "panel_messages": [
                        {"panel_id": "A", "message": "Panel A must not claim external validation."}
                    ],
                    "threshold_semantics": "Thresholds are descriptive operating points.",
                    "recommendation_boundary": "This figure should not be framed as treatment guidance.",
                }
            ],
        },
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(
        encoding="utf-8"
    )
    assert "Caption." in submission_markdown
    assert "The cohort architecture is summarized for reader orientation." in submission_markdown
    assert submission_markdown.count("The cohort architecture is summarized for reader orientation.") == 1
    assert "The figure summarizes the observed cohort architecture." not in submission_markdown
    assert "Thresholds are descriptive operating points." not in submission_markdown
    assert "paper should" not in submission_markdown.lower()
    assert "do not recast" not in submission_markdown.lower()
    assert "must not" not in submission_markdown.lower()
    assert "should not" not in submission_markdown.lower()


def test_figure_legend_merge_skips_near_duplicate_semantic_message() -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal_parts.markdown_surface")

    legend = module.merge_legend_with_figure_semantics(
        base_legend=(
            "Recorded metabolic diagnostic fields by BMI category among records with populated status. "
            "Point size reflects the available denominator; labels show positive records over available records. "
            "Percentages exclude unknown values and are not prevalence estimates."
        ),
        figure_semantics={
            "direct_message": (
                "Recorded metabolic diagnostic fields are shown by BMI category among records with populated status. "
                "Point size reflects the available denominator; labels show positive records over available records. "
                "Percentages exclude unknown values and are not prevalence estimates."
            )
        },
    )

    assert legend.count("Recorded metabolic diagnostic fields") == 1


def test_create_submission_minimal_package_skips_stale_panel_messages_for_single_panel_grouped_calibration(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_paper_workspace(tmp_path)

    dump_json(
        paper_root / "figure_semantics_manifest.json",
        {
            "schema_version": 1,
            "figures": [
                {
                    "figure_id": "F1",
                    "direct_message": "Observed mortality rises steeply across NHANES deciles while mean predicted risk remains compressed.",
                    "panel_messages": [
                        {"panel_id": "A", "message": "Panel A: legacy multi-panel wording should not survive the single-panel grouped calibration export."},
                        {"panel_id": "B", "message": "Panel B: legacy multi-panel wording should not survive the single-panel grouped calibration export."},
                    ],
                }
            ],
        },
    )

    figure_catalog_path = paper_root / "figures" / "figure_catalog.json"
    figure_catalog = json.loads(figure_catalog_path.read_text(encoding="utf-8"))
    figure_catalog["figures"][0]["template_id"] = "fenggaolab.org.medical-display-core::time_to_event_risk_group_summary"
    figure_catalog["figures"][0]["qc_result"] = {
        "status": "pass",
        "metrics": {"plot_variant": "nhanes_decile_grouped_calibration"},
    }
    dump_json(figure_catalog_path, figure_catalog)

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(
        encoding="utf-8"
    )
    assert "Observed mortality rises steeply across NHANES deciles while mean predicted risk remains compressed." in submission_markdown
    assert "legacy multi-panel wording should not survive the single-panel grouped calibration export" not in submission_markdown


def test_create_submission_minimal_package_copies_deferred_supplementary_figure_assets(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_paper_workspace(tmp_path)

    figure_catalog_path = paper_root / "figures" / "figure_catalog.json"
    figure_catalog = json.loads(figure_catalog_path.read_text(encoding="utf-8"))
    figure_catalog["figures"] = [figure_catalog["figures"][0]]
    figure_catalog["deferred_figures"] = [
        {
            "figure_id": "F4",
            "paper_role": "supplementary",
            "display_role": "deferred_context_not_main_evidence",
            "title": "Deferred threshold-utility context",
            "caption": "Deferred context retained as supplementary support.",
        }
    ]
    dump_json(figure_catalog_path, figure_catalog)
    output_png_path = paper_root / "figures" / "generated" / "F4_time_to_event_decision_curve.png"
    output_pdf_path = paper_root / "figures" / "generated" / "F4_time_to_event_decision_curve.pdf"
    write_png(output_png_path)
    write_text(output_pdf_path, "%PDF-1.4\n%deferred figure\n")
    dump_json(
        paper_root / "build" / "display_pack_render_requests" / "F4.render_request.json",
        {
            "output_png_path": str(output_png_path.resolve()),
            "output_pdf_path": str(output_pdf_path.resolve()),
            "output_svg_path": None,
        },
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    figure_names = {path.name for path in (paper_root / "submission_minimal" / "figures").iterdir()}
    assert "Figure1.png" in figure_names
    assert "Figure4.png" in figure_names
    assert "Figure4.pdf" in figure_names


def test_create_submission_minimal_package_materializes_supplementary_tables_workbook(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_paper_workspace(tmp_path)
    write_text(
        paper_root / "tables" / "S1_audit_dictionary.md",
        """# Supplementary Table S1

| paper_role | analysis_denominator | claim_boundary |
| --- | --- | --- |
| supplementary | eligible indicator denominator | recorded care-review gap only |
""",
    )
    write_text(
        paper_root / "tables" / "S1_audit_dictionary.csv",
        "paper_role,analysis_denominator,claim_boundary\nsupplementary,eligible indicator denominator,recorded care-review gap only\n",
    )
    table_catalog_path = paper_root / "tables" / "table_catalog.json"
    table_catalog = json.loads(table_catalog_path.read_text(encoding="utf-8"))
    table_catalog["tables"].append(
        {
            "table_id": "S1",
            "paper_role": "supplementary",
            "title": "Supplementary audit dictionary",
            "asset_paths": [
                "paper/tables/S1_audit_dictionary.csv",
                "paper/tables/S1_audit_dictionary.md",
            ],
        }
    )
    dump_json(table_catalog_path, table_catalog)

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    workbook_path = paper_root / "submission_minimal" / "supplementary_tables.xlsx"
    assert workbook_path.exists()
    assert manifest["supplementary_material"]["tables_workbook_path"] == (
        "paper/submission_minimal/supplementary_tables.xlsx"
    )

    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path)
    assert "S1" in workbook.sheetnames
    sheet = workbook["S1"]
    assert sheet["A1"].value == "Supplementary Table S1. Supplementary audit dictionary"
    assert [cell.value for cell in sheet[3][:3]] == [
        "paper_role",
        "analysis_denominator",
        "claim_boundary",
    ]
    assert [cell.value for cell in sheet[4][:3]] == [
        "supplementary",
        "eligible indicator denominator",
        "recorded care-review gap only",
    ]


def test_create_submission_minimal_package_preserves_top_level_figures_in_manuscript_shaped_draft(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    draft_path = paper_root / "draft.md"

    write_png(paper_root / "figures" / "F1.png")
    write_text(
        draft_path,
        draft_path.read_text(encoding="utf-8")
        + """

# Figures

## Figure 1. Preserved top-level figure

![](figures/F1.png)

Preserved legend for the manuscript-shaped top-level figures block.
""",
    )

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_root = paper_root / "submission_minimal"
    submission_markdown = (submission_root / "manuscript_submission.md").read_text(encoding="utf-8")

    assert submission_markdown.splitlines().count("# Main Figures") == 1
    assert "## Figure 1. Preserved top-level figure" in submission_markdown
    assert "![](figures/Figure1.png)" in submission_markdown
    assert "Preserved legend for the manuscript-shaped top-level figures block." in submission_markdown

    manuscript_surface_qc = manifest["manuscript"]["surface_qc"]
    assert manuscript_surface_qc["status"] == "pass"
    assert manuscript_surface_qc["source_markdown"]["figure_blocks_with_images"] == 1
    assert manuscript_surface_qc["source_markdown"]["figure_blocks_with_legends"] == 1


def test_create_submission_minimal_package_strips_short_figure_id_from_image_alt_text(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    draft_path = paper_root / "draft.md"

    write_png(paper_root / "figures" / "F1.png")
    write_text(
        draft_path,
        draft_path.read_text(encoding="utf-8")
        + """

# Figures

## F1. Main figure

![F1. Main figure](figures/F1.png)

Legend text for the short-F main figure.
""",
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(
        encoding="utf-8"
    )
    assert "## Figure 1. Main figure" in submission_markdown
    assert "![Main figure](figures/Figure1.png){width=100%}" in submission_markdown
    assert "![F1. Main figure]" not in submission_markdown


def test_create_submission_minimal_package_orders_main_figures_by_figure_number(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    draft_path = paper_root / "draft.md"
    write_png(paper_root / "figures" / "F2_main.png")
    write_png(paper_root / "figures" / "F3_main.png")
    dump_json(
        paper_root / "figures" / "figure_catalog.json",
        {
            "schema_version": 1,
            "figures": [
                {
                    "figure_id": "F2",
                    "paper_role": "main_text",
                    "title": "Second figure",
                    "export_paths": ["paper/figures/F2_main.png"],
                },
                {
                    "figure_id": "F3",
                    "paper_role": "main_text",
                    "title": "Third figure",
                    "export_paths": ["paper/figures/F3_main.png"],
                },
                {
                    "figure_id": "F1",
                    "paper_role": "main_text",
                    "title": "First figure",
                    "export_paths": ["paper/figures/F1_main.png"],
                },
            ],
        },
    )
    write_text(
        draft_path,
        draft_path.read_text(encoding="utf-8")
        + """

# Figures

## Figure 2. Second figure

![](figures/F2_main.png)

## Figure 3. Third figure

![](figures/F3_main.png)

## Figure 1. First figure

![](figures/F1_main.png)
""",
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
        route_context=_open_submission_route_context(),
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(encoding="utf-8")
    figure_headings = re.findall(r"^## (Figure \d+\.[^\n]+)", submission_markdown, flags=re.MULTILINE)
    assert figure_headings == [
        "Figure 1. First figure",
        "Figure 2. Second figure",
        "Figure 3. Third figure",
    ]


def test_create_submission_minimal_package_splits_nested_tables_and_orders_by_table_number(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    draft_path = paper_root / "draft.md"
    write_text(
        draft_path,
        draft_path.read_text(encoding="utf-8")
        + """

# Main Tables

## Table 1

### Table 3. Third support table

| Metric | Value |
| --- | --- |
| C | 3 |

### Table 1. First support table

| Metric | Value |
| --- | --- |
| A | 1 |

### Table 2. Second wide support table

| First long header | Second long header | Third long header | Fourth long header | Fifth long header | Sixth long header |
| --- | --- | --- | --- | --- | --- |
| A | B | C | D | E | F |
""",
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
        route_context=_open_submission_route_context(),
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(encoding="utf-8")
    table_headings = re.findall(r"^## (Table \d+\.[^\n]+)", submission_markdown, flags=re.MULTILINE)
    assert table_headings == [
        "Table 1. First support table",
        "Table 2. Second wide support table",
        "Table 3. Third support table",
    ]
    wide_separator = re.search(r"\| First long header .+\|\n\| ([^\n]+) \|", submission_markdown)
    assert wide_separator is not None
    separator_widths = [len(cell.strip()) for cell in wide_separator.group(1).split("|")]
    assert separator_widths == [24, 18, 17, 18, 17, 17]


def test_create_submission_minimal_package_renders_extra_wide_tables_as_sideways_latex(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    draft_path = paper_root / "draft.md"
    write_text(
        draft_path,
        draft_path.read_text(encoding="utf-8")
        + """

# Main Tables

## Table 1. Extra-wide support table

| Phenotype | Index patients | Share | Mean age | Mean BMI | Mean HbA1c | Severe glycemia gap | Uncontrolled glycemia no drug | Hypertension no antihypertensive |
| --- | --- | --- | --- | --- | --- | --- | --- | --- |
| Cardiometabolic-risk dominant diabetes | 138797 | 20.04% | 68.86 | 21.91 | 5.74 | NA | NA | 60.48% |
""",
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
        route_context=_open_submission_route_context(),
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(encoding="utf-8")
    assert "\\begin{landscape}" in submission_markdown
    assert "\\caption*{Table 1. Extra-wide support table}" in submission_markdown
    assert "\\begin{tabular}{p{0.20\\linewidth}" in submission_markdown
    assert "Cardiometabolic-risk dominant diabetes & 138797 & 20.04\\%" in submission_markdown


def test_create_submission_minimal_package_uses_catalog_markdown_for_long_measure_value_tables(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    draft_path = paper_root / "draft.md"
    write_text(
        draft_path,
        draft_path.read_text(encoding="utf-8")
        + """

# Main Tables

## Table 2. Baseline characteristics and recorded treatment-review gaps by phenotype

| Phenotype | Index patients | Share of index cohort | Mean age, y | Mean BMI | Mean HbA1c | Severe glycemia low-intensity gap | Uncontrolled glycemia with no diabetes drug | Hypertension with no antihypertensive | Dyslipidemia with no lipid-lowering |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
| Glycemic-dominant diabetes | 104029 | 15.02% | 64.44 | 23.05 | 8.04 | 86.11% | 50.05% | 71.52% | 85.55% |
""",
    )
    write_text(
        paper_root / "tables" / "generated" / "T2_phenotype_gap_summary.md",
        """# Phenotype-level clinical characteristics and treatment-gap rates

| Phenotype | Measure | Value |
| --- | --- | --- |
| Glycemic-dominant diabetes | Index patients | 104029 |
| Glycemic-dominant diabetes | Severe glycemia low-intensity gap | 86.11% |
""",
    )
    dump_json(
        paper_root / "tables" / "table_catalog.json",
        {
            "schema_version": 1,
            "tables": [
                {
                    "table_id": "T2",
                    "paper_role": "main_text",
                    "title": "Baseline characteristics and recorded treatment-review gaps by phenotype",
                    "asset_paths": [
                        "paper/tables/generated/T2_phenotype_gap_summary.csv",
                        "paper/tables/generated/T2_phenotype_gap_summary.md",
                    ],
                    "render_result": {
                        "table_layout_policy": "long_measure_value_table_to_avoid_pdf_header_overlap",
                    },
                }
            ],
        },
    )
    write_text(
        paper_root / "tables" / "generated" / "T2_phenotype_gap_summary.csv",
        "Phenotype,Measure,Value\nGlycemic-dominant diabetes,Index patients,104029\n",
    )

    module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
        route_context=_open_submission_route_context(),
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(encoding="utf-8")
    assert "| Phenotype | Measure | Value |" in submission_markdown
    assert "Severe glycemia low-intensity gap | Uncontrolled glycemia" not in submission_markdown
    assert "## Table 2. Baseline characteristics and recorded treatment-review gaps by phenotype" in submission_markdown


def test_create_submission_minimal_package_accepts_materialized_submission_source_from_compile_report(
    tmp_path: Path,
    real_submission_exports,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_materialized_submission_source_workspace(tmp_path)

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_root = paper_root / "submission_minimal"
    assert (submission_root / "manuscript_source.md").exists()
    submission_markdown = (submission_root / "manuscript_submission.md").read_text(encoding="utf-8")
    assert 'title: "Materialized Submission Title"' in submission_markdown
    assert "# Main Figures" in submission_markdown
    assert "## Figure 1. Main figure" in submission_markdown
    assert "![F1](figures/Figure1.png){width=100%}" in submission_markdown
    assert "Materialized figure caption." in submission_markdown
    manuscript_surface_qc = manifest["manuscript"]["surface_qc"]
    assert manuscript_surface_qc["status"] == "pass"
    assert manifest["manuscript"]["source_markdown_alias_role"] == "authority_note"
    assert manuscript_surface_qc["authority_note"]["role_clarity_pass"] is True

    with zipfile.ZipFile(submission_root / "manuscript.docx") as archive:
        names = archive.namelist()
        document_xml = archive.read("word/document.xml").decode("utf-8", errors="ignore")
    assert any(name.startswith("word/media/") for name in names)
    assert "<w:drawing" in document_xml

    pdf_reader = PdfReader(str(submission_root / "paper.pdf"))
    assert sum(len(page.images) for page in pdf_reader.pages) >= 1


def test_describe_submission_minimal_authority_accepts_materialized_submission_source_from_compile_report(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_materialized_submission_source_workspace(tmp_path)

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    authority = module.describe_submission_minimal_authority(paper_root=paper_root)

    assert authority["status"] == "current"
    assert authority["stale_reason"] is None
    assert authority["recorded_source_signature"] == manifest["source_signature"]
    assert authority["source_signature"] == manifest["source_signature"]


def test_create_submission_minimal_package_falls_back_when_compile_report_points_to_missing_submission_source(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_current_draft_workspace(tmp_path)

    dump_json(
        paper_root / "build" / "compile_report.json",
        {
            "source_markdown_path": "paper/submission_minimal/manuscript_source.md",
            "pdf_path": "paper/paper.pdf",
        },
    )

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_root = paper_root / "submission_minimal"
    submission_markdown = (submission_root / "manuscript_submission.md").read_text(encoding="utf-8")
    assert 'title: "Current Draft Title"' in submission_markdown
    assert manifest["manuscript"]["source_markdown_path"] == "paper/submission_minimal/manuscript_submission.md"


def test_create_submission_minimal_package_materializes_references_and_pending_front_matter(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_current_draft_workspace(tmp_path)

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_root = paper_root / "submission_minimal"
    copied_references_path = submission_root / "references.bib"
    assert copied_references_path.exists()
    assert copied_references_path.read_text(encoding="utf-8") == (
        paper_root / "references.bib"
    ).read_text(encoding="utf-8")

    assert manifest["references"] == {
        "source_path": "paper/references.bib",
        "source_kind": "paper_references",
        "output_path": "paper/submission_minimal/references.bib",
        "entry_count": 1,
        "coverage": {
            "status": "complete",
            "citation_key_count": 1,
            "missing_citation_keys": [],
        },
    }
    assert manifest["front_matter_placeholders"] == {
        "authors": "pending",
        "affiliations": "pending",
        "corresponding_author": "pending",
        "funding": "pending",
        "conflict_of_interest": "pending",
        "ethics": "pending",
        "data_availability": "pending",
        "analytic_data_lock_date": "pending",
        "registry_enrollment_period": "pending",
    }


def test_create_submission_minimal_package_uses_workspace_literature_references_when_paper_refs_missing(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_current_draft_workspace(tmp_path)
    workspace_root = paper_root.parent
    (paper_root / "references.bib").unlink()
    workspace_references_path = workspace_root / "memory" / "portfolio" / "research_memory" / "literature" / "references.bib"
    write_text(
        workspace_references_path,
        """@article{ref1,
  title={A workspace primary source},
  author={Author, A. and Author, B.},
  journal={Journal},
  year={2024}
}
""",
    )

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_root = paper_root / "submission_minimal"
    submission_markdown = (submission_root / "manuscript_submission.md").read_text(encoding="utf-8")
    copied_references_path = submission_root / "references.bib"
    assert "bibliography: references.bib" in submission_markdown
    assert copied_references_path.read_text(encoding="utf-8") == workspace_references_path.read_text(
        encoding="utf-8"
    )
    assert manifest["references"]["source_path"] == "memory/portfolio/research_memory/literature/references.bib"
    assert manifest["references"]["source_kind"] == "workspace_literature"
    assert manifest["references"]["coverage"] == {
        "status": "complete",
        "citation_key_count": 1,
        "missing_citation_keys": [],
    }


def test_general_medical_submission_renders_bibliography_when_refs_exist_without_inline_citations(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    paper_root = make_manuscript_shaped_draft_workspace(tmp_path)
    draft_text = (paper_root / "draft.md").read_text(encoding="utf-8")
    write_text(
        paper_root / "draft.md",
        re.sub(r"\s*\[@ref1\]", "", draft_text),
    )

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    submission_markdown = (paper_root / "submission_minimal" / "manuscript_submission.md").read_text(
        encoding="utf-8"
    )
    assert "bibliography: references.bib" in submission_markdown
    assert "nocite: '@*'" in submission_markdown
    assert "\n# References\n" in submission_markdown
    assert manifest["references"]["coverage"] == {
        "status": "nocite_all",
        "citation_key_count": 0,
        "bib_entry_count": 1,
        "missing_citation_keys": [],
        "rendered_bibliography_policy": "pandoc_nocite_all",
    }


def test_create_submission_minimal_package_auto_hydrates_missing_pubmed_references(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    shared_base = importlib.import_module("med_autoscience.controllers.submission_minimal_parts.shared_base")
    paper_root = make_current_draft_workspace(tmp_path)
    write_text(
        paper_root / "draft.md",
        (paper_root / "draft.md").read_text(encoding="utf-8").replace("[@ref1]", "[@pmid_12345]"),
    )
    write_text(paper_root / "references.bib", "")
    fetched_record = LiteratureRecord(
        record_id="pmid:12345",
        title="PubMed hydrated evidence",
        authors=("A. Author",),
        year=2025,
        journal="Lancet",
        doi="10.1000/pubmed",
        pmid="12345",
        pmcid=None,
        arxiv_id=None,
        abstract=None,
        full_text_availability="abstract_only",
        source_priority=2,
        citation_payload={"source": "test"},
        local_asset_paths=(),
        relevance_role="candidate",
        claim_support_scope=(),
    )
    synced_records: list[dict[str, object]] = []

    monkeypatch.setattr(
        shared_base.pubmed_adapter,
        "fetch_pubmed_summary",
        lambda *, pmids: [fetched_record] if pmids == ["12345"] else [],
    )
    monkeypatch.setattr(
        shared_base.workspace_literature_controller,
        "sync_workspace_literature",
        lambda *, workspace_root, records: synced_records.extend(records) or {"status": "synchronized"},
    )

    manifest = module.create_submission_minimal_package(
        paper_root=paper_root,
        publication_profile="general_medical_journal",
    )

    paper_references_text = (paper_root / "references.bib").read_text(encoding="utf-8")
    submission_references_text = (paper_root / "submission_minimal" / "references.bib").read_text(
        encoding="utf-8"
    )
    assert "@article{pmid_12345," in paper_references_text
    assert "title = {PubMed hydrated evidence}" in submission_references_text
    assert manifest["references"]["coverage"]["status"] == "complete"
    assert manifest["references"]["coverage"]["auto_repair"]["status"] == "repaired"
    assert manifest["references"]["coverage"]["auto_repair"]["fetched_pmids"] == ["12345"]
    assert manifest["references"]["coverage"]["auto_repair"]["workspace_literature_sync"] == {
        "status": "synchronized"
    }
    assert synced_records == [asdict(fetched_record)]


def test_create_submission_minimal_package_rejects_unrepairable_missing_reference_keys(
    tmp_path: Path,
) -> None:
    module = importlib.import_module("med_autoscience.controllers.submission_minimal")
    shared_base = importlib.import_module("med_autoscience.controllers.submission_minimal_parts.shared_base")
    paper_root = make_current_draft_workspace(tmp_path)
    write_text(
        paper_root / "draft.md",
        (paper_root / "draft.md").read_text(encoding="utf-8").replace("[@ref1]", "[@missing_key]"),
    )
    write_text(paper_root / "references.bib", "")

    with pytest.raises(shared_base.SubmissionReferenceCoverageError, match="missing_key"):
        module.create_submission_minimal_package(
            paper_root=paper_root,
            publication_profile="general_medical_journal",
        )
